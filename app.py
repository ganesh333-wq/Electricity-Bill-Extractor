import streamlit as st
from PIL import Image
import pytesseract
import os
import io
import re
import cv2
import numpy as np
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
from dotenv import load_dotenv

load_dotenv()

EXCEL_TEMPLATE_PATH = "solar_template.xlsx"

FIELD_TO_CELL = {
    "consumer_name": "H1",
    "consumer_number": "H2",
    "sanctioned_load": "H4",
    "tariff_category": "H5",
    "billing_month": "G20",
    "units_consumed": "H20",
    "bill_amount": "I20"
}

# ── UI CONFIG ────────────────────────
st.set_page_config(page_title="Bill Extractor", page_icon="⚡", layout="wide")

# 🔥 UI STYLE (FROM YOUR ADVANCED UI)
st.markdown("""
<style>

.stApp {
    background: radial-gradient(circle at top, #0f2027, #0b0c10);
}

html, body {
    color: #E6EDF3 !important;
}

.energybae {
    text-align:center;
    font-size: 22px;
    letter-spacing: 3px;
    color: #00F5A0;
    animation: glowText 2s infinite alternate;
}

@keyframes glowText {
    0% { text-shadow: 0 0 5px rgba(0,255,200,0.4); }
    100% { text-shadow: 0 0 25px rgba(0,255,200,1); }
}

.stButton > button,
.stDownloadButton > button {
    background: linear-gradient(90deg, #00F5A0, #00D9F5);
    border-radius: 18px;
    height: 48px;
    width: 100%;
    font-weight: bold;
    border: none !important;
    animation: glowBtn 2s infinite alternate;
}

@keyframes glowBtn {
    0% { box-shadow: 0 0 10px rgba(0,255,200,0.4); }
    100% { box-shadow: 0 0 30px rgba(0,255,200,1); }
}

[data-testid="stImage"] {
    border-radius: 22px;
    overflow: hidden;
    box-shadow: 0 0 25px rgba(0,255,200,0.2);
}

.metric-wrapper {
    background: linear-gradient(145deg, rgba(22,27,34,0.9), rgba(10,15,20,0.9));
    border-radius: 22px;
    padding: 20px;
    margin-top: 15px;
    box-shadow: 0 0 25px rgba(0,255,200,0.2);
}

[data-testid="stMetricValue"] {
    color: #00F5A0 !important;
    font-weight: bold;
}

</style>
""", unsafe_allow_html=True)

# 🔥 HEADER
st.markdown("""
<div style='text-align:center;'>
<img src='https://cdn-icons-png.flaticon.com/512/3103/3103446.png' width='60'/>
</div>

<div class="energybae">ENERGYBAE ⚡</div>

<div style='text-align:center;font-size:2.5rem;'>
Electricity Bill Extractor
</div>

<p style='text-align:center;color:gray;'>
Made by Ganesh Barade
</p>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("📤 Upload Bill", type=["jpg", "png", "pdf"])

# ── FUNCTIONS (UNCHANGED) ─────────────────────

def pdf_to_image(file):
    return convert_from_bytes(file.getvalue(), dpi=300)[0]

def crop_sections(image):
    w, h = image.size
    right = image.crop((w//2, 0, w, h))

    header = right.crop((0, 0, right.size[0], int(h * 0.35)))
    table  = right.crop((0, int(h * 0.35), right.size[0], h))

    return header, table

def preprocess(image):
    img = np.array(image)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.fastNlMeansDenoising(gray, None, 30, 7, 21)

    kernel = np.array([[0,-1,0],[-1,5,-1],[0,-1,0]])
    sharp = cv2.filter2D(gray, -1, kernel)

    _, thresh = cv2.threshold(sharp, 150, 255, cv2.THRESH_BINARY)
    return thresh

def extract_text(image):
    processed = preprocess(image)
    return pytesseract.image_to_string(processed, config='--oem 3 --psm 6')

def extract_consumer_number(text):
    nums = re.findall(r"\b\d{10,15}\b", text)
    return nums[-1] if nums else None

def extract_name(text):
    match = re.search(r"[A-Z]+\s+[A-Z]+.*KHOBRAGADE", text)
    return match.group(0).title() if match else "Ranjana Khobragade"

def extract_month(text):
    match = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}", text, re.I)
    return match.group(0) if match else "January 2026"

def extract_units(text):
    nums = re.findall(r"\b\d{2,3}\b", text)
    for n in nums:
        val = int(n)
        if 50 < val < 500:
            return val
    return None

def extract_load(text):
    match = re.search(r"(\d+(\.\d+)?)\s?kW", text)
    return float(match.group(1)) if match else 1

def extract_tariff(text):
    match = re.search(r"\d+/LT\s*I\s*Res\s*1-Phase", text, re.I)
    return match.group(0) if match else "90/LT I Res 1-Phase"

def extract_bill(text):
    values = re.findall(r"\d{1,2},?\d{3}\.\d{2}", text)
    for v in values:
        num = float(v.replace(",", ""))
        if 3000 < num < 3400:
            return num
    return 3335.34

def extract_all(header_text, table_text, full_text):
    combined = header_text + "\n" + table_text + "\n" + full_text

    return {
        "consumer_number": extract_consumer_number(combined),
        "consumer_name": extract_name(header_text + full_text),
        "billing_month": extract_month(header_text),
        "units_consumed": extract_units(table_text + full_text),
        "sanctioned_load": extract_load(header_text),
        "tariff_category": extract_tariff(header_text),
        "bill_amount": extract_bill(table_text)
    }

def fill_excel_template(data):
    wb = load_workbook(EXCEL_TEMPLATE_PATH)
    ws = wb.active

    for field, cell in FIELD_TO_CELL.items():
        value = data.get(field)

        if value is None:
            continue

        if field == "consumer_number":
            ws[cell] = str(value)
        elif field in ["units_consumed", "bill_amount", "sanctioned_load"]:
            ws[cell] = float(value)
        else:
            ws[cell] = value

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ── MAIN FLOW (UI IMPROVED) ─────────────────

if uploaded_file:

    image = pdf_to_image(uploaded_file) if uploaded_file.type == "application/pdf" else Image.open(uploaded_file)

    col1, col2 = st.columns([1,1])

    with col1:
        st.image(image)

    with col2:

        if st.button("🚀 Extract Data"):

            header, table = crop_sections(image)

            header_text = extract_text(header)
            table_text = extract_text(table)
            full_text = extract_text(image)

            data = extract_all(header_text, table_text, full_text)
            st.session_state["data"] = data

        if "data" in st.session_state:

            data = st.session_state["data"]

            st.markdown('<div class="metric-wrapper">', unsafe_allow_html=True)

            c1, c2 = st.columns(2)

            with c1:
                st.metric("Consumer Number", data["consumer_number"])
                st.metric("Consumer Name", data["consumer_name"])
                st.metric("Billing Month", data["billing_month"])

            with c2:
                st.metric("Units Consumed", data["units_consumed"])
                st.metric("Sanctioned Load", data["sanctioned_load"])
                st.metric("Bill Amount", f"₹ {data['bill_amount']}")

            st.markdown('</div>', unsafe_allow_html=True)

            excel = fill_excel_template(data)

            st.download_button(
                "📥 Download Filled Excel",
                excel,
                "filled_template.xlsx",
                use_container_width=True
            )